
import os, re, sys, graphing, openpyxl
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle
from operator import itemgetter

#(Date)(, )(Hours:Minutes)(Seconds)(AM/PM)(:/ -)( )(Sender)(: )(Message)
messageFormat = re.compile(r'^([0-9]{1,2}/[0-9]{2}/[0-9]{2})(, )([0-9]{1,2}:[0-9]{2})(:[0-9]{2})*( [A-Z]{2})( -|:)( )(.*?)(: )(.*)')
#[Hours]:[Minutes]
timeSplit = re.compile(r'([0-9]{1,2}):([0-9]{2})')

#To get rid of file extension when making graphing
fileSplit = re.compile(r'(.*)(.[a-zA-Z0-9]{3,4})')

#Initializing empty lists of dictionaries for date, time, person and word, no of messages to 0 and a message list
noMessages = 0
dateDictionary = []
timeDictionary = []
personDictionary = []
wordDictionary = []
messageList = []

#Copy contents of file to a variable.
def readFile(fileName):
    fi = open(fileName, 'r', encoding='utf-8')
    textToAnalyze = fi.readlines()
    fi.close()
    return textToAnalyze

#If value exists in the dictionary, increment count. Else add a new entry.
def analyze(dictionary, value, keyName):
    for d in dictionary:
        if d[keyName] == value:
            d['Count'] += 1
            break
    else:
        dictionary.append({keyName: value, 'Count': 1})
    return dictionary

#Sort a dictionary in descending order of its count
def sort(dictionary):
    return sorted(dictionary, key=itemgetter('Count'), reverse=True)

#Get a dictionary of the most used words in the chat and the number of times they're used while ignoring commonly used words.
def getWordFrequency(messageList):
    wordFile = open('commonWords.txt', 'r')
    commonWordsList = wordFile.read()
    wordFile.close()
    stripChars = re.compile(r'[a-zA-z0-9]+')
    frequencyList = []
    for messages in messageList:
        wordsList = messages.split(' ')
        for words in wordsList:
            words = words.lower()
            if words in commonWordsList:
                continue
            if stripChars.search(words):
                words = stripChars.search(words).group()
                for existingWords in frequencyList:
                    if existingWords['Word'] == words:
                        existingWords['Count'] += 1
                        break
                else:
                    frequencyList.append({'Word': words, 'Count': 1})
    return sort(frequencyList)

#save to excel sheet
def toXL(dictionary, sheetName, col1, col2):
    #If file exists, add a new sheet
    if os.path.isfile("data.xlsx"):
        xl = openpyxl.load_workbook("data.xlsx")
        xl.create_sheet(title=sheetName)
        sheet = xl.get_sheet_by_name(sheetName)
    #If not use current sheet
    else:
        xl = openpyxl.Workbook()
        sheet = xl.get_active_sheet()
        sheet.title = sheetName

    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 15
    #Setup headings
    headings = [col1, col2]
    rowNum = 2
    for colNo, heading in enumerate(headings):
        sheet.cell(row=rowNum, column=colNo + 3).value = heading

    rowNum += 1

    #Add data
    for item in dictionary:
        sheet.cell(row=rowNum, column=3).value = item[col1]
        sheet.cell(row=rowNum, column=4).value = item['Count']
        rowNum += 1
    xl.save("data.xlsx")

#Gets file name from command line arguments
def getFileName():
    #If no filename is given
    if len(sys.argv) < 2:
        print('Usage: analyze.py fileName')
        sys.exit()
    #Get file name from command line
    return ' '.join(sys.argv[1:])

fileName = getFileName()

#Analyze file
textToAnalyze = readFile(fileName)

#for each line, check if it is a new message. If it is, split it up into components.
#The date and sender components are added to their dictionaries.
#Split up the time into hours, minutes, seconds and AM/PM.
#If it is PM, add 12 hours to it. If hours is a single digit, convert it to double digits.
#Each message is appended to a list for later analysis.
for lines in textToAnalyze:
    if messageFormat.search(lines):
        found = messageFormat.search(lines)
        dateDictionary = analyze(dateDictionary, found[1], 'Date')
        personDictionary = analyze(personDictionary, found[8], 'Sender')
        #Time
        time = timeSplit.search(found[3])
        hours = time[1]
        if len(hours) == 1:
            hours = '0' + str(hours)
        if found[5] == ' PM':
            if (int(hours) + 12) >= 24:
                hours = 0
            else:
                hours = int(hours) + 12
        timeDictionary = analyze(timeDictionary, str(hours), 'Time')
        #Message. Ignore media.
        # -*- coding: utf-8 -*-
        if "<Media omitted>" not in found[10] and "<‎attached>" not in found[10]:
            messageList.append(found[10])
        noMessages += 1

#Sort dictionaries and get word dictionary from analysing the message list
dateDictionary = sort(dateDictionary)
personDictionary = sort(personDictionary)
timeDictionary = sorted(timeDictionary, key=itemgetter('Time'))
wordDictionary = getWordFrequency(messageList)

#remove old data sheets
if os.path.isfile("data.xlsx"):
    os.unlink("data.xlsx")

 #Add to excel sheet
toXL(dateDictionary, 'Dates', 'Date', 'No. of Messages')
toXL(personDictionary, 'People', 'Sender', 'No. of Messages')
toXL(timeDictionary, 'Times', 'Time', 'No. of Messages')
toXL(wordDictionary, 'Words', 'Word', 'No. of Uses')

#Generate graphing
newFileName = fileSplit.search(fileName)[1]
graphing.histogram(timeDictionary, 'Message Time Chart in ' + newFileName, 'timeActivity.png')
graphing.barGraph(wordDictionary[:15], 'Word', 'Uses', 'Most used words in ' + str(noMessages) + ' messages in ' + newFileName, 'wordFrequency.png')
graphing.barGraph(dateDictionary[:15], 'Date', 'Messages', 'Most Messages in ' + newFileName, 'dateActivity.png')
graphing.barGraph(personDictionary[:15], 'Sender', 'Messages', 'Most active person in ' + newFileName, 'personActivity.png')
